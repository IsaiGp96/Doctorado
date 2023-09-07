from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook


app = Flask(__name__)


@app.route('/psomain', methods=['POST'])
def index():
    # Parámetros de control (ingresan)
    w = float(request.form['w'])
    n = 9  # Es una constante en tu código, si quieres que sea variable, puedes cambiarlo.
    c1 = float(request.form['c1'])
    c2 = float(request.form['c2'])
    e = int(request.form['e'])
    itera = e  #iteraciones

#Función objetivo
# Ri= (t^-i)/ (t^+i+t^-i) , i=1,...,m
# vector negativo / (vector positivo + vector negarivo)
# R2/(R2+R1)


    # Variables aleatorias, estas van a ser ingresadas(posteriormente)
    #R1 --> es el Vector positivo
    FILE_PATH='r1.xlsx'
    SHEET='Hoja1'
    workbook=load_workbook(FILE_PATH,read_only=True)
    sheet=workbook[SHEET]
    r1=[]
    for row in sheet.iter_rows():
        #print(row[0].value)
        r1.append(row[0].value)
        #print(r1)

    #R2 --> es el vector negativo
    #SHEET = 'Hoja1'
    FILE_PATH='r2.xlsx'
    workbook=load_workbook(FILE_PATH,read_only=True)
    sheet=workbook[SHEET]
    r2=[]
    for row in sheet.iter_rows():
        #print(row[0].value)
        r2.append(row[0].value)
        #print(r1)

    # ----------
    # Inicialización de la partícula del enjambre
    # Iniciando cálculo de las iteraciones
    print("")
    print("----")
    print("Iteración # 1")
    CP=[]
    for i in range(n):
        r11=float(r1[i])
        CP.append(float(format(10*(r11-0.5), '.4f')))
        #print("CP=",CP)
    #print("   CP=",CP)
    print("   Posición actual:       CP=",CP)

    # Inicialización de la velocidad
    V=[]
    for i in range(n):
        r21=float(r2[i])
        V.append(float(format((r21-0.5), '.4f')))
    #print("   V=",V)
    #print("   Velocidad actual:             V=",V)

    # Óptimo actual CF(1)
    # Ri= (t^-i)/ (t^+i+t^-i) , i=1,...,m
    # vector negativo / (vector positivo + vector negarivo)
    # R2/(R2+R1)
    CF=[]
    for i in range(n):
        #R2/(R1+R2)
        CForm1=float(r1[i])
        CForm2=float(r2[i])
        CForm3=(float((CForm2)/(CForm2+CForm1)))
        CForm=(format(float(CForm3), '.4f'))
        CF.append(float(CForm))
    print("   Óptimo actual:         CF=",CF)
    #print("   CF=",CF)

    # mejor posición actual
    LBP=[]
    for i in range(n):
        # LBP[1]=CP[1]
        LBPP=format(float(CP[i]), '.4f')
        LBP.append(float(LBPP))
    #print("   LBP=",LBP)
    print("   Mejor posición local:  LBP=",LBP)

    # mejor óptimo local
    LBF=[]
    for i in range(n):
        # LBF[1]=CF[1]
        LBFF=format(float(CF[i]), '.4f')
        LBF.append(float(LBFF))
    #print("   LBF=",LBF)
    print("   Mejor óptimo local:    LBF=",LBF)

    # MEJOR ÓPTIMO GLOBAL
    GBF=[]
    GBF.append(max(LBF))
    GBFt=max(LBF)
    #print("   GBF=", GBF)
    print("   Mejor óptimo global:   GBF=", GBF)

    # MEJOR POSICION GLOBAL
    GBP=[]
    GBF_index=LBF.index(GBFt)
    GBPt=float(LBP[GBF_index])
    GBP.append(LBP[GBF_index])
    #print("   GBP=",GBP)
    print("   Mejor posición global: GBP=",GBP)

    ###########################################
    #  Iteráción 2 a la n
    ##########################################
    it=2
    e=e-2

    while (e>=0):
        print("--------")
        print("Iteración #", it)
        long_V1=len(V)-n
        long_LBP1=len(LBP)-n
        long_CP1=len(CP)-n
        long_GBP1=len(GBP)-1
        for i in range(n):
            # V(i+1)= W*V(i) +c1*r1*(LBP(i)-CP(i))+c2*r2*(GBP)i)-CP(i))
            V1=format(float(w)*float(V[long_V1]), '.4f')
            V2=format(float(c1)*float(r1[i]), '.4f')
            V3=format(float(LBP[long_LBP1])-float(CP[long_CP1]), '.4f')
            V4=format(float(c2)*float(r2[i]), '.4f')
            V5=format(float(GBP[long_GBP1]), '.4f')
            V6=format(float(CP[long_CP1]), '.4f')
            V7=float(V5)-float(V6)
            Vx=format((float(V1)+float(V2)*float(V3)+float(V4)*float(V7)), '.4f')
            V.append(Vx)
            long_V1=long_V1+1
            long_LBP1=long_LBP1+1
            long_CP1=long_CP1+1
        long_V2=len(V)-n
        long_CP2=len(CP)-n

        #CP=CP(i)+V(i)
        for i in range(n):
            CPI=(format(float(CP[long_CP2])+float(V[long_V2]), '.4f'))
            CP.append(float(CPI))
            long_V2=long_V2+1
            long_CP2=long_CP2+1

        #óptimo actual CF
        long_CP1=len(CP)-(2*n)
        #print("long_CP2",long_CP2)
        long_CP2=len(CP)-n
        #print("long_CP3",long_CP3)
        for i in range(n):
            #CF2/(CF2+CF1)
            #print("CP",CP)
            CF1=CP[long_CP1]
            #print("CF1",CF1)
            CF2=CP[long_CP2]
            #print("CF2",CF2)
            CF3=float(CF2+CF1)
            CF12=format(float((CF2)/(CF3)),'.4f')
            CF.append(CF12)
            #print("CF12 ",CF12)
            long_CP1=long_CP1+1
            long_CP2=long_CP2+1

        # mejor óptimo local
        long_CF2=len(CF)-n
        long_LBF1=len(LBF)-(n)
        for i in range(n):
            #Max( CF[i],LBF[i-1])
            #print("CF",CF)
            #print("LBF",LBF)

            CFt=float(CF[long_CF2])
            #print("long_CF2",long_CF2)
            #print("CFt",CFt)
            
            LBFt=float(LBF[long_LBF1])
            #print("long_LBF1",long_LBF1)
            #print("LBFt",LBFt)
            if CFt>LBFt:
                LBF.append(CFt)
            else:
                LBF.append(LBFt)
            #print("LBF",LBF)
            long_CF2=long_CF2+1
            long_LBF1=long_LBF1+1

        # mejor posición actual
        long_CP4=len(CP)-n
        for i in range(n):
            #LBP[i]= posición de LBF[i]-CP[i]
            LBPt=(format(float(CP[long_CP4]), '.4f'))
            LBP.append(float(LBPt))
            long_CP4=long_CP4+1

        # MEJOR ÓPTIMO GLOBAL
        long_LBF=len(CP)-n
        temporal=[]
        for i in range(n):
            temporal.append(LBF[long_LBF])
            long_LBF=long_LBF+1
        GBF.append(max(temporal))
        GBFt=max(temporal)
        #print("GBFt=",GBFt)
        #print("MaxGlobal=",GBFt)

        # MEJOR POSICION GLOBAL
        #print("PosiciónGBP posición en CP")
        #print("CP=",CP)
        #long_GBF=len(CP)-n
        long_GBF=len(CP)-1
        #print("LBF",LBF)
        #print("long_CP=",long_GBF)
        GBPt=CP[long_GBF]
        #print("GBPt",GBPt)
        GBP.append(GBPt)


    # impresión de datos por iteración
        V_imp = len(V)-n
        CP_imp = len(CP)-n
        CF_imp = len(CF)-n
        LBF_imp = len(LBF)-n
        LBP_imp = len(LBP)-n
        GBF_imp = len(GBF)-1
        GBP_imp = len(GBP)-1

        V_impT = []
        for i in range(n):
            V_impT.append(float(V[V_imp]))
            V_imp = V_imp+1
        print("   V=",V_impT)

        CP_impT = []
        for i in range(n):
            CP_impT.append(float(CP[CP_imp]))
            CP_imp = CP_imp+1
        print("   CP=",CP_impT)

        CF_impT = []
        for i in range(n):
            CF_impT.append(float(CF[CF_imp]))
            CF_imp = CF_imp+1
        print("   CF=",CF_impT)

        LBF_impT = []
        for i in range(n):
            LBF_impT.append(float(LBF[LBF_imp]))
            LBF_imp = LBF_imp+1
        print("   LBF=",LBF_impT)

        LBP_impT = []
        for i in range(n):
            LBP_impT.append(float(LBP[LBP_imp]))
            LBP_imp = LBP_imp+1
        print("   LBP=",LBP_impT)

        for i in range(n):
            GBF_impT = (float(GBF[GBF_imp]))
        print("   GBF=",GBF_impT)

        for i in range(n):
            GBP_impT = (float(GBP[GBP_imp]))
        print("   GBP=",GBP_impT)

        e = e-1
        it = it+1
        print("")

    print("*******************")
    GBF_SF = len(GBF)-1
    GBP_SF = len(GBP)-1

    GBF_FIN = (float(GBF[GBF_SF]))
    print("   Mejor posición=",GBF_FIN)

    GBP_FIN = (float(GBP[GBP_SF]))
    print("   Mejor óptimo=",GBP_FIN)
    print("*******************")
    print("")

    context = {
        "c1": c1,
        "c2": c2,
        "w": w,
        "n": n,
        "e": e,
        "itera": itera,
        #"var1": var1,
        #"var2": var2,
        #"var3": var3,

        "r1": r1,
        "r2": r2,

        "V_impT": V_impT,
        "CP_impT": CP_impT,
        "CF_impT": CF_impT,
        "LBF_impT": LBF_impT,
        "LBP_impT": LBP_impT,
        "GBF_impT": GBF_impT,
        "GBP_impT": GBP_impT,

        "GBF_FIN": GBF_FIN,
        "GBP_FIN": GBP_FIN,
    }

    return jsonify(context)


@app.route('/')
def calcular():
    # Aquí también puedes retornar un archivo HTML o algún mensaje.
    return "psomain"


if '__main__' == __name__:
    app.run(port=5000, debug=True)
